#!/usr/bin/env python3
"""
scripts/parse-new-master.py

Parse the original (DD Mau "started it all") inventory xlsx into
scripts/data/new-master.json. The JSON is what scripts/mapper-server.mjs
loads as the target/master list for the matching workflow.

Why Python (not Node): the repo doesn't have the xlsx npm package as a
dependency, and adding it just for a once-or-twice parser was overkill.
Python ships openpyxl by default in most macOS Python installs.

Usage:
    python3 scripts/parse-new-master.py [path/to/xlsx]
    # default path is ~/Downloads/Inventory 9.12.25.xlsx

Notes on the xlsx structure (DD Mau's original inventory worksheet):
  • Two sheets: "Daily schedule" (shorter copy) and "Sheet2" (full).
  • Columns: A=Spanish name, B=English name, C=TUES marker, D=FRI/SAT,
    E=Vendor option 1, F=Vendor option 2, G=QTY (count).
  • Category headers are rows where col C == "TUES" — the original
    worksheet uses a fresh column-header row to separate sections.
    The "Proteins" group at the top has no header (it's implicit).
  • Vendor abbreviations are normalized to the canonical names used in
    src/data/inventory.js (WS → STL Wholesale, RD → Restaurant Depot,
    J → Jays, etc.).
"""

import json
import sys
import datetime
import os
from collections import OrderedDict

DEFAULT_XLSX = os.path.expanduser("~/Downloads/Inventory 9.12.25.xlsx")
OUT_PATH = os.path.join(os.path.dirname(__file__), "data", "new-master.json")

VENDOR_MAP = {
    "WS": "STL Wholesale",
    "RD": "Restaurant Depot",
    "RD?": "Restaurant Depot",
    "RD OR US FOODS": "Restaurant Depot",
    "US FOODS": "US Foods",
    "USFOODS": "US Foods",
    "COSTCO": "Costco Business",
    "SYSCO": "Sysco",
    "J": "Jays",
    "JAYS": "Jays",
    "FABULOUS FISH": "Fabulous Fish",
    "PAN-ASIA": "Pan Asia",
    "PANASIA": "Pan Asia",
    "WING-HING": "Wing Hing",
    "WINGHING": "Wing Hing",
    "WHOLE FOODS": "Whole Foods",
    "WF": "Whole Foods",
    "YE": "Yellowstone",
    "YEI": "Yellowstone",
    "OLIVE MARKET": "Olive Market",
    "WEBSTAURANT": "Webstaurant",
    "SPECIAL ORDER": "Special Order",
}

def canon_vendor(s):
    if not s:
        return ""
    key = s.strip().upper().replace("  ", " ")
    if key in VENDOR_MAP:
        return VENDOR_MAP[key]
    return s.strip().title()


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(src):
        print(f"ERROR: file not found: {src}", file=sys.stderr)
        sys.exit(1)

    from openpyxl import load_workbook
    wb = load_workbook(src, data_only=True)
    ws = wb["Sheet2"]

    items = []
    current_cat = "Proteins"  # implicit first category before "VEGGIES"
    for r in range(2, ws.max_row + 1):
        es = ws.cell(row=r, column=1).value
        en = ws.cell(row=r, column=2).value
        tues = ws.cell(row=r, column=3).value
        fri = ws.cell(row=r, column=4).value
        v1 = ws.cell(row=r, column=5).value
        v2 = ws.cell(row=r, column=6).value

        # Header row patterns (the xlsx isn't consistent — some headers
        # also include the column header strip "TUES / FRI / OPTION 1
        # / OPTION 2", others are just the bare section name).
        # We accept a row as a header if either:
        #   1. col C contains "TUES" (the full column-header row), or
        #   2. col A and col B are non-empty, identical, all-uppercase
        #      strings (so "VEGGIES" / "CLEANING SUPPLIES" qualify but
        #      "Salmon" / "Ginger" don't — they're real items whose
        #      es and en names happen to be the same).
        if tues and str(tues).strip().upper() in ("TUES", "TUESDAY"):
            current_cat = (en or "").strip()
            continue
        if en and es and str(en).strip() == str(es).strip():
            s = str(en).strip()
            if s and s == s.upper() and any(c.isalpha() for c in s):
                current_cat = s
                continue
        if not en and not es:
            continue
        items.append({
            "cat": current_cat,
            "en": (en or "").strip() if en else "",
            "es": (es or "").strip() if es else "",
            "tues": True if tues and "X" in str(tues).upper() else False,
            "fri": True if fri and "X" in str(fri).upper() else False,
            "v1": canon_vendor(v1),
            "v2": canon_vendor(v2),
        })

    by_cat = OrderedDict()
    for it in items:
        by_cat.setdefault(it["cat"], []).append(it)

    snapshot = {
        "sourceFile": os.path.basename(src),
        "parsedAt": datetime.datetime.now().isoformat(),
        "totalItems": len(items),
        "categories": [
            {
                "name": cat,
                "items": [
                    {
                        "en": it["en"],
                        "es": it["es"],
                        "tues": it["tues"],
                        "fri": it["fri"],
                        "vendor1": it["v1"],
                        "vendor2": it["v2"],
                    }
                    for it in lst
                ],
            }
            for cat, lst in by_cat.items()
        ],
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w") as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False)

    print(f"Wrote {OUT_PATH}")
    print(f"  source     : {snapshot['sourceFile']}")
    print(f"  items      : {snapshot['totalItems']}")
    print(f"  categories : {len(snapshot['categories'])}")
    for cat in snapshot["categories"]:
        print(f"    - {cat['name']:<25} {len(cat['items'])} items")


if __name__ == "__main__":
    main()
